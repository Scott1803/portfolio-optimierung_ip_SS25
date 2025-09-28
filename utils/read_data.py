from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List

import numpy as np
from openpyxl import load_workbook

# Relative to the repository root (adjust if needed)
EXCEL_FILE_RELATIVE_PATH = "../data/auswertung_input.xlsx"


# =============================
# Data classes
# =============================
@dataclass
class SocialImpactOptimizationInputData:
    """
    Data used to calculate the social impact for an investment portfolio.

    - `projects_budget_min_max`: m = 2 x n = [project count] sized matrix, with the first row
      containing the minimum investable amounts in each project, and the second row containing
      the maximum investable amounts
    - `country_indicator_scores_weighted`: matrix of size m = [indicator count] x (1 + [country count]).
      Rows are indicators; columns are the base weight value [column 0] and the countries' weighted
      indicator scores [columns 1..].
    - `project_country_weights`: m = [project count] x n = [country count] matrix containing the weight
      scores of the projects per country (how much project x affects country y)
    - `sia_indikatoren_verbesserungen`: m = [indicator count] x n = [project count] matrix containing
      the percentages each project can maximally improve each indicator (relative improvement!)
    """
    projects_budget_min_max: np.ndarray
    country_indicator_scores_weighted: np.ndarray
    project_country_weights: np.ndarray
    sia_indikatoren_verbesserungen: np.ndarray


@dataclass
class EconomicImpactOptimizationInputData:
    projects_budget_min_max: np.ndarray
    ia_sektorenwerte: List[np.ndarray]
    ia_projekt_budget_aufteilung: np.ndarray
    ia_laender_aufteilung: np.ndarray
    ia_projekt_abhaengigkeiten: np.ndarray

    """
    Data used to calculate the economic impact for an investment portfolio.

    - `projects_budget_min_max`: 2 x n = [project count] matrix with min/max investable amounts
    - `ia_sektorenwerte`: For each country (order given by the sheet), a matrix of size
      [sector count] x 2 where column 0 is "v_a" and column 1 is "o_mult".
    - `ia_projekt_budget_aufteilung`: [sector count] x [project count] percentage split of project budgets per sector
    - `ia_laender_aufteilung`: [country count] x [project count] describing how project budgets are split between countries
    - `ia_projekt_abhaengigkeiten`: [project count] x [project count] matrix with 0/1 indicating dependencies
      (rows are dependencies of the columns; e.g., if [11][3] == 1, then project 4 is a dependency of project 12)
    """


@dataclass
class IdMappingInputData:
    """Dictionaries where keys are the IDs used in the data tables and values are human readable names."""
    indicators: Dict[str, str]
    projects: Dict[str, str]
    sectors: Dict[str, str]
    countries: Dict[str, str]


# =============================
# Utilities
# =============================

def _resolve_excel_path() -> Path:
    """Resolve the Excel file path relative to this file's directory.

    Raises
    ------
    FileNotFoundError
        If the file cannot be found at the resolved path.
    """
    base_dir = Path(__file__).resolve().parent
    path = (base_dir / EXCEL_FILE_RELATIVE_PATH).resolve()
    if not path.exists():
        # Also allow absolute path in the constant for flexibility
        alt = Path(EXCEL_FILE_RELATIVE_PATH)
        if alt.is_absolute() and alt.exists():
            return alt
        raise FileNotFoundError(
            f"Excel input file not found at '{path}'. Set EXCEL_FILE_RELATIVE_PATH appropriately."
        )
    return path


def _to_float(val) -> float:
    """Coerce a cell value to float, handling commas as decimal separators and blanks as 0.

    Examples
    --------
    "1,25" -> 1.25, "-0,5" -> -0.5, "" or None -> 0.0
    """
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s == "" or s.lower() in {"nan", "none"}:
        return 0.0
    # replace thousands separators if any, then decimal comma with dot
    s = s.replace(" ", "").replace("\xa0", "")
    # Some sheets might contain percent-style strings; we do *not* divide by 100 automatically.
    # Caller should ensure semantics. Here we only parse numerically.
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError as e:
        raise ValueError(f"Cannot convert value '{val}' to float") from e


def _require_sheet(wb, name: str):
    if name not in wb.sheetnames:
        raise KeyError(f"Missing required sheet '{name}' in Excel file")
    return wb[name]


# =============================
# Readers
# =============================

def read_social_impact_optimization_input_data() -> SocialImpactOptimizationInputData:
    """
    Attempts to read tables from the input data excel sheet:
    - `projekte_budget_min_max`
    - `sia_indikatoren_laender_scores`
    - `sia_projekte_laender_gewichte`
    - `sia_indikatoren_verbesserungen`

    Returns
    -------
    SocialImpactOptimizationInputData

    Raises
    ------
    FileNotFoundError
        If the Excel file cannot be found
    KeyError
        If any required sheet is missing
    ValueError
        If the sheet structure is unexpected
    """
    path = _resolve_excel_path()
    wb = load_workbook(path, data_only=True)

    # --- projekte_budget_min_max ---
    ws_bud = _require_sheet(wb, "projekte_budget_min_max")
    # Expect header with project IDs in row 1, values in rows 2 (min) and 3 (max)
    header = [c.value for c in ws_bud[1]]
    # Project IDs start at column 2
    proj_ids = [v for v in header[1:] if v is not None and str(v).strip() != ""]
    if len(ws_bud[2]) < 2 or len(ws_bud[3]) < 2:
        raise ValueError("projekte_budget_min_max: Expected at least 3 rows (header, min, max)")
    mins = [_to_float(c.value) for c in ws_bud[2][1:1 + len(proj_ids)]]
    maxs = [_to_float(c.value) for c in ws_bud[3][1:1 + len(proj_ids)]]
    projects_budget_min_max = np.vstack([mins, maxs]).astype(float)

    # --- sia_indikatoren_laender_scores ---
    ws_scores = _require_sheet(wb, "sia_indikatoren_laender_scores")
    # Header: [Indicator, Weights, <country codes...>]
    header_scores = [c.value for c in ws_scores[1]]
    if len(header_scores) < 3 or str(header_scores[0]).lower() != "indicator":
        raise ValueError("sia_indikatoren_laender_scores: Unexpected header; expected first cell 'Indicator'")
    # Country columns (after 'Weights')
    # column 0: indicator id, column 1: Weights, columns 2..: per-country weighted values
    data_rows = []
    for row in ws_scores.iter_rows(min_row=2, max_row=ws_scores.max_row, values_only=True):
        if row[0] is None or str(row[0]).strip() == "":
            continue  # skip blank rows
        weight = _to_float(row[1])
        country_vals = [_to_float(v) for v in row[2:]]
        data_rows.append([weight] + country_vals)
    country_indicator_scores_weighted = np.array(data_rows, dtype=float)

    # --- sia_projekte_laender_gewichte ---
    ws_pcw = _require_sheet(wb, "sia_projekte_laender_gewichte")
    # Header: ["", AL, RO/RU, ...]
    header_pcw = [c.value for c in ws_pcw[1]]
    # countries begin from column 2 (index 1)
    country_cols = [v for v in header_pcw[1:] if v is not None and str(v).strip() != ""]
    pcw_rows = []
    for row in ws_pcw.iter_rows(min_row=2, max_row=ws_pcw.max_row, values_only=True):
        if row[0] is None or str(row[0]).strip() == "":
            continue
        vals = [_to_float(v) for v in row[1:1 + len(country_cols)]]
        pcw_rows.append(vals)
    project_country_weights = np.array(pcw_rows, dtype=float)

    # --- sia_indikatoren_verbesserungen ---
    ws_improve = _require_sheet(wb, "sia_indikatoren_verbesserungen")
    header_impr = [c.value for c in ws_improve[1]]
    project_cols = [v for v in header_impr[1:] if v is not None and str(v).strip() != ""]
    impr_rows = []
    for row in ws_improve.iter_rows(min_row=2, max_row=ws_improve.max_row, values_only=True):
        if row[0] is None or str(row[0]).strip() == "":
            continue
        vals = [_to_float(v) for v in row[1:1 + len(project_cols)]]
        impr_rows.append(vals)
    sia_indikatoren_verbesserungen = np.array(impr_rows, dtype=float)

    return SocialImpactOptimizationInputData(
        projects_budget_min_max=projects_budget_min_max,
        country_indicator_scores_weighted=country_indicator_scores_weighted,
        project_country_weights=project_country_weights,
        sia_indikatoren_verbesserungen=sia_indikatoren_verbesserungen,
    )


def read_economic_impact_optimization_input_data() -> EconomicImpactOptimizationInputData:
    """
    Attempts to read tables from the input data excel sheet:
    - `projekte_budget_min_max`
    - `ia_sektorenwerte`
    - `ia_projekt_budget_aufteilung`
    - `ia_laender_aufteilung`
    - `ia_projekt_abhaengigkeiten`

    Returns
    -------
    EconomicImpactOptimizationInputData

    Raises
    ------
    FileNotFoundError
        If the Excel file cannot be found
    KeyError
        If any required sheet is missing
    ValueError
        If the sheet structure is unexpected
    """
    path = _resolve_excel_path()
    wb = load_workbook(path, data_only=True)

    # --- projekte_budget_min_max --- (reuse logic)
    ws_bud = _require_sheet(wb, "projekte_budget_min_max")
    header = [c.value for c in ws_bud[1]]
    proj_ids = [v for v in header[1:] if v is not None and str(v).strip() != ""]
    if len(ws_bud[2]) < 2 or len(ws_bud[3]) < 2:
        raise ValueError("projekte_budget_min_max: Expected at least 3 rows (header, min, max)")
    mins = [_to_float(c.value) for c in ws_bud[2][1:1 + len(proj_ids)]]
    maxs = [_to_float(c.value) for c in ws_bud[3][1:1 + len(proj_ids)]]
    projects_budget_min_max = np.vstack([mins, maxs]).astype(float)

    # --- ia_sektorenwerte ---
    ws_sek = _require_sheet(wb, "ia_sektorenwerte")
    # Expect two header rows: Row 1 with country codes repeated every two columns, Row 2 with [v_a, o_mult] headers
    r1 = [c.value for c in ws_sek[1]]
    r2 = [c.value for c in ws_sek[2]]
    if len(r1) < 2 or len(r2) < 2:
        raise ValueError("ia_sektorenwerte: Unexpected header structure")

    # Build pairs of columns per country
    pairs: List[tuple[str, int]] = []  # [(country_code, start_col_index_0_based), ...]
    col = 0
    while col < len(r1):
        country = r1[col]
        # Column titles might include blanks; ensure next column exists
        if country is not None and str(country).strip() != "":
            # The next column should be the second half of the pair
            if col + 1 >= len(r2) or str(r2[col]).strip().lower() not in {"v_a", "v a", "va"}:
                # It's acceptable that row 2 labels shift; we won't be strict about exact text
                pass
            pairs.append((str(country).strip(), col))
            col += 2
        else:
            col += 1

    if not pairs:
        raise ValueError("ia_sektorenwerte: No country column pairs detected in header")

    # Read sector rows (starting from row 3) for each country pair
    ia_sektorenwerte: List[np.ndarray] = []
    for country, start_idx in pairs:
        rows = []
        for row in ws_sek.iter_rows(min_row=3, max_row=ws_sek.max_row, values_only=True):
            # Protect against ragged rows
            v_a = _to_float(row[start_idx]) if start_idx < len(row) else 0.0
            o_mult = _to_float(row[start_idx + 1]) if (start_idx + 1) < len(row) else 0.0
            rows.append([v_a, o_mult])
        ia_sektorenwerte.append(np.array(rows, dtype=float))

    # --- ia_projekt_budget_aufteilung ---
    ws_pba = _require_sheet(wb, "ia_projekt_budget_aufteilung")
    header_pba = [c.value for c in ws_pba[1]]
    project_cols = [v for v in header_pba[1:] if v is not None and str(v).strip() != ""]
    pba_rows = []
    for row in ws_pba.iter_rows(min_row=2, max_row=ws_pba.max_row, values_only=True):
        if row[0] is None or str(row[0]).strip() == "":
            continue
        vals = [_to_float(v) for v in row[1:1 + len(project_cols)]]
        pba_rows.append(vals)
    ia_projekt_budget_aufteilung = np.array(pba_rows, dtype=float)

    # --- ia_laender_aufteilung ---
    ws_cpa = _require_sheet(wb, "ia_laender_aufteilung")
    header_cpa = [c.value for c in ws_cpa[1]]
    project_cols2 = [v for v in header_cpa[1:] if v is not None and str(v).strip() != ""]
    cpa_rows = []
    for row in ws_cpa.iter_rows(min_row=2, max_row=ws_cpa.max_row, values_only=True):
        if row[0] is None or str(row[0]).strip() == "":
            continue
        vals = [_to_float(v) for v in row[1:1 + len(project_cols2)]]
        cpa_rows.append(vals)
    ia_laender_aufteilung = np.array(cpa_rows, dtype=float)

    # --- ia_projekt_abhaengigkeiten ---
    ws_dep = _require_sheet(wb, "ia_projekt_abhaengigkeiten")
    header_dep = [c.value for c in ws_dep[1]]
    project_cols3 = [v for v in header_dep[1:] if v is not None and str(v).strip() != ""]
    dep_rows = []
    for row in ws_dep.iter_rows(min_row=2, max_row=ws_dep.max_row, values_only=True):
        if row[0] is None or str(row[0]).strip() == "":
            continue
        vals = [int(round(_to_float(v))) for v in row[1:1 + len(project_cols3)]]
        dep_rows.append(vals)
    ia_projekt_abhaengigkeiten = np.array(dep_rows, dtype=int)

    return EconomicImpactOptimizationInputData(
        projects_budget_min_max=projects_budget_min_max,
        ia_sektorenwerte=ia_sektorenwerte,
        ia_projekt_budget_aufteilung=ia_projekt_budget_aufteilung,
        ia_laender_aufteilung=ia_laender_aufteilung,
        ia_projekt_abhaengigkeiten=ia_projekt_abhaengigkeiten,
    )


def read_id_mapping_input_data() -> IdMappingInputData:
    """
    Attempts to read the `mappings` sheet and parse the following dictionaries:
    - indicators: {I1: name, I2: name, ...}
    - projects: {P1: name, P2: name, ...}
    - sectors: {S1: name, S2: name, ...}
    - countries: {<code>: name, ...}  (e.g., AL, RO/RU, SE, MO, BU)

    Returns
    -------
    IdMappingInputData

    Raises
    ------
    FileNotFoundError
        If the Excel file cannot be found
    KeyError
        If the sheet is missing
    """
    path = _resolve_excel_path()
    wb = load_workbook(path, data_only=True)
    ws = _require_sheet(wb, "mappings")

    # The sheet is structured as repeated pairs of columns labeled 'ID' and 'Name'.
    header = [c.value for c in ws[1]]
    pairs: List[int] = []  # list of starting column indices for each (ID, Name) pair
    col = 0
    while col < len(header) - 1:
        if (str(header[col]).strip().lower() == "id") and (str(header[col + 1]).strip().lower() == "name"):
            pairs.append(col)
            col += 2
        else:
            col += 1

    if not pairs:
        raise ValueError("mappings: Could not detect any (ID, Name) header pairs")

    indicators: Dict[str, str] = {}
    projects: Dict[str, str] = {}
    sectors: Dict[str, str] = {}
    countries: Dict[str, str] = {}

    for start_col in pairs:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            _id = row[start_col] if start_col < len(row) else None
            _name = row[start_col + 1] if (start_col + 1) < len(row) else None
            if _id is None or str(_id).strip() == "":
                continue
            key = str(_id).strip()
            val = "" if _name is None else str(_name).strip()

            if key.upper().startswith("I"):
                indicators[key] = val
            elif key.upper().startswith("P"):
                projects[key] = val
            elif key.upper().startswith("S"):
                sectors[key] = val
            else:
                # Assume country code (e.g., AL, RO/RU, MO, BU, SE)
                countries[key] = val

    return IdMappingInputData(
        indicators=indicators,
        projects=projects,
        sectors=sectors,
        countries=countries,
    )
